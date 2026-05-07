from flask import Flask, render_template, request, redirect, url_for, flash, session
from functools import wraps
import sqlite3
import os
import shutil
from datetime import datetime
import openpyxl

app = Flask(__name__)
app.secret_key = 'claro-fortinet-2026'
DB_PATH = os.path.join(os.path.dirname(__file__), 'sistema.db')
BACKUP_DIR = os.path.join(os.path.dirname(__file__), 'backups')
ALLOW_DB_RESET = os.environ.get('ALLOW_DB_RESET', '').lower() in ('1', 'true', 'yes', 'on')

USUARIO = 'EstratOpera'
SENHA   = 'Bolsao26'


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logado'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def init_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.execute("PRAGMA journal_mode=WAL")
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS pontos_bolsao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            point_pack_number TEXT NOT NULL UNIQUE,
            responsavel TEXT NOT NULL,
            projetos TEXT,
            pontos INTEGER NOT NULL,
            used_amount REAL DEFAULT 0,
            registration_date TEXT,
            expiration_date TEXT
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS pontos_utilizados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bolsao_id INTEGER,
            serial_number TEXT NOT NULL,
            dados_cliente TEXT,
            product_model TEXT,
            valor_pontos_dia REAL NOT NULL,
            data_aplicacao TEXT NOT NULL,
            data_fim TEXT,
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS base_conciliacao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            serial_number TEXT NOT NULL,
            description TEXT,
            usage_date TEXT,
            points REAL NOT NULL
        )
    ''')
    conn.commit()
    conn.close()


init_db()


def get_db_connection():
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    return conn


def backup_database(label='snapshot'):
    if not os.path.exists(DB_PATH):
        return None

    os.makedirs(BACKUP_DIR, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'sistema_{label}_{timestamp}.db'
    backup_path = os.path.join(BACKUP_DIR, filename)
    shutil.copy2(DB_PATH, backup_path)
    return backup_path


# ── Autenticação ──────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    erro = None
    if request.method == 'POST':
        if request.form['usuario'] == USUARIO and request.form['senha'] == SENHA:
            session['logado'] = True
            return redirect(url_for('dashboard'))
        erro = 'Usuário ou senha incorretos.'
    return render_template('login.html', erro=erro)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    conn = get_db_connection()

    bolsao_summary = conn.execute('''
        SELECT
            responsavel || ' (' || projetos || ')' as grupo,
            SUM(pontos) as pontos_totais,
            SUM(used_amount) as used_totais_fortinet
        FROM pontos_bolsao
        GROUP BY grupo
    ''').fetchall()

    utilizados_summary = conn.execute('''
        SELECT
            b.responsavel || ' (' || b.projetos || ')' as grupo,
            SUM(pu.valor_pontos_dia * (julianday('now') - julianday(pu.data_aplicacao))) as pontos_consumidos_calculado
        FROM pontos_utilizados pu
        JOIN pontos_bolsao b ON pu.bolsao_id = b.id
        GROUP BY grupo
    ''').fetchall()

    conn.close()

    utilizados_map = {r['grupo']: r['pontos_consumidos_calculado'] for r in utilizados_summary}
    dados_dashboard = []

    for item in bolsao_summary:
        grupo               = item['grupo']
        pontos_totais       = item['pontos_totais']
        used_fortinet       = item['used_totais_fortinet']
        pontos_calc         = utilizados_map.get(grupo, 0)
        dados_dashboard.append({
            'grupo':                    grupo,
            'pontos_totais':            pontos_totais,
            'used_totais_fortinet':     used_fortinet,
            'remaining_totais_fortinet': pontos_totais - used_fortinet,
            'pontos_utilizados_analitico': pontos_calc,
            'faltantes_analitico':      pontos_totais - pontos_calc,
            'percent_fortinet':         (used_fortinet / pontos_totais * 100) if pontos_totais else 0,
            'percent_analitico':        (pontos_calc   / pontos_totais * 100) if pontos_totais else 0,
        })

    return render_template('index.html', dashboard_data=dados_dashboard)


# ── Pontos Bolsão ─────────────────────────────────────────────────────────────

@app.route('/pontos_bolsao')
@login_required
def listar_pontos_bolsao():
    conn = get_db_connection()
    pontos = conn.execute('SELECT * FROM pontos_bolsao ORDER BY registration_date DESC').fetchall()
    conn.close()
    return render_template('pontos_bolsao.html', pontos=pontos)


@app.route('/pontos_bolsao/novo', methods=['GET', 'POST'])
@login_required
def novo_ponto_bolsao():
    erro = None
    if request.method == 'POST':
        try:
            conn = get_db_connection()
            conn.execute('''
                INSERT INTO pontos_bolsao
                    (point_pack_number, responsavel, projetos, pontos, used_amount, registration_date, expiration_date)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                request.form['point_pack_number'],
                request.form['responsavel'],
                request.form['projetos'],
                int(request.form['pontos']),
                float(request.form.get('used_amount') or 0),
                request.form['registration_date'],
                request.form['expiration_date'],
            ))
            conn.commit()
            conn.close()
            backup_database('pontos_bolsao')
            return redirect(url_for('listar_pontos_bolsao'))
        except sqlite3.IntegrityError:
            erro = 'Número do Pack já cadastrado. Use um número diferente.'
        except Exception as e:
            erro = f'Erro ao salvar: {str(e)}'

    return render_template('novo_bolsao.html', erro=erro)

# ── Pontos Utilizados ─────────────────────────────────────────────────────────

@app.route('/pontos_utilizados')
@login_required
def listar_pontos_utilizados():
    conn = get_db_connection()
    pontos = conn.execute('''
        SELECT
            pu.id, pu.serial_number, pu.dados_cliente, pu.product_model,
            pu.valor_pontos_dia, pu.data_aplicacao, pu.data_fim,
            b.responsavel || ' (' || b.projetos || ')' as resp_projeto,
            CASE
                WHEN pu.data_fim IS NULL OR pu.data_fim >= date('now')
                THEN CAST(julianday('now') - julianday(pu.data_aplicacao) AS INTEGER)
                ELSE CAST(julianday(pu.data_fim) - julianday(pu.data_aplicacao) AS INTEGER)
            END as dias_consumidos,
            CASE
                WHEN pu.data_fim IS NULL OR pu.data_fim >= date('now')
                THEN CAST(julianday('now') - julianday(pu.data_aplicacao) AS INTEGER)
                ELSE CAST(julianday(pu.data_fim) - julianday(pu.data_aplicacao) AS INTEGER)
            END * pu.valor_pontos_dia as pontos_consumidos
        FROM pontos_utilizados pu
        JOIN pontos_bolsao b ON pu.bolsao_id = b.id
        ORDER BY pu.data_aplicacao DESC
    ''').fetchall()
    conn.close()

    total_pontos = sum(p['pontos_consumidos'] for p in pontos)
    media_pontos = total_pontos / len(pontos) if pontos else 0
    return render_template('pontos_utilizados.html', data=pontos,
                           total_pontos=total_pontos, media_pontos=media_pontos)


@app.route('/pontos_utilizados/novo', methods=['GET', 'POST'])
@login_required
def novo_ponto_utilizado():
    if request.method == 'POST':
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO pontos_utilizados
                (bolsao_id, serial_number, dados_cliente, product_model, valor_pontos_dia, data_aplicacao, data_fim)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            request.form['bolsao_id'],
            request.form['serial_number'],
            request.form['dados_cliente'],
            request.form['product_model'],
            float(request.form['valor_pontos_dia']),
            request.form['data_aplicacao'],
            request.form.get('data_fim') or None,
        ))
        conn.commit()
        conn.close()
        backup_database('pontos_utilizados')
        return redirect(url_for('listar_pontos_utilizados'))

    conn = get_db_connection()
    bolsoes = conn.execute('''
        SELECT id, responsavel, projetos, (pontos - used_amount) as saldo
        FROM pontos_bolsao
        ORDER BY responsavel, projetos, saldo DESC
    ''').fetchall()
    conn.close()

    grupos_vistos, grupos_unicos = set(), []
    for b in bolsoes:
        chave = (b['responsavel'], b['projetos'])
        if chave not in grupos_vistos:
            grupos_vistos.add(chave)
            grupos_unicos.append(b)

    return render_template('novo_ponto_utilizado.html', bolsoes=grupos_unicos)


# ── Conciliação ───────────────────────────────────────────────────────────────

@app.route('/conciliacao', methods=['GET', 'POST'])
@login_required
def conciliacao():
    conn = get_db_connection()

    if request.method == 'POST':
        arquivo = request.files.get('arquivo_conciliacao')
        if not arquivo or arquivo.filename == '':
            flash('Nenhum arquivo selecionado.', 'erro')
            return redirect(url_for('conciliacao'))

        if os.path.splitext(arquivo.filename)[1].lower() not in ('.xlsx', '.xls'):
            flash('Formato inválido. Envie um arquivo .xlsx ou .xls.', 'erro')
            return redirect(url_for('conciliacao'))

        try:
            wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

            def col_idx(name):
                for i, h in enumerate(headers):
                    if name.lower() in h:
                        return i
                return None

            idx_serial = col_idx('serial')
            idx_desc   = col_idx('description')
            idx_date   = col_idx('usage date') or col_idx('date')
            idx_points = col_idx('points')

            if idx_serial is None or idx_points is None:
                flash('Colunas obrigatórias não encontradas. O arquivo deve conter "Serial Number" e "Points".', 'erro')
                return redirect(url_for('conciliacao'))

            conn.execute('DELETE FROM base_conciliacao')
            rows_inseridos = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                serial = row[idx_serial]
                if not serial:
                    continue
                desc   = row[idx_desc]   if idx_desc   is not None else None
                date   = row[idx_date]   if idx_date   is not None else None
                points = row[idx_points] if idx_points is not None else 0

                if isinstance(date, datetime):
                    date = date.strftime('%Y-%m-%d')
                elif date is not None:
                    date = str(date)

                try:
                    points = float(points) if points is not None else 0.0
                except (ValueError, TypeError):
                    points = 0.0

                conn.execute(
                    'INSERT INTO base_conciliacao (serial_number, description, usage_date, points) VALUES (?, ?, ?, ?)',
                    (str(serial).strip(), str(desc).strip() if desc else '', date, points)
                )
                rows_inseridos += 1

            conn.commit()
            flash(f'Base importada com sucesso! {rows_inseridos} registros carregados.', 'sucesso')
            backup_database('conciliacao')

        except Exception as e:
            flash(f'Erro ao processar o arquivo: {str(e)}', 'erro')
        finally:
            conn.close()

        return redirect(url_for('conciliacao'))

    resultado = conn.execute('''
        SELECT
            pu.serial_number, pu.dados_cliente, pu.product_model,
            b.responsavel || ' (' || b.projetos || ')' AS grupo,
            pu.valor_pontos_dia, pu.data_aplicacao,
            CASE
                WHEN pu.data_fim IS NULL OR pu.data_fim >= date('now')
                THEN CAST(julianday('now') - julianday(pu.data_aplicacao) AS INTEGER)
                ELSE CAST(julianday(pu.data_fim) - julianday(pu.data_aplicacao) AS INTEGER)
            END AS dias_consumidos,
            CASE
                WHEN pu.data_fim IS NULL OR pu.data_fim >= date('now')
                THEN CAST(julianday('now') - julianday(pu.data_aplicacao) AS INTEGER)
                ELSE CAST(julianday(pu.data_fim) - julianday(pu.data_aplicacao) AS INTEGER)
            END * pu.valor_pontos_dia AS pontos_calculados,
            COALESCE((
                SELECT SUM(bc.points) FROM base_conciliacao bc
                WHERE UPPER(TRIM(bc.serial_number)) = UPPER(TRIM(pu.serial_number))
            ), 0) AS pontos_fortinet
        FROM pontos_utilizados pu
        JOIN pontos_bolsao b ON pu.bolsao_id = b.id
        ORDER BY grupo, pu.serial_number
    ''').fetchall()

    total_base = conn.execute('SELECT COUNT(*) as c FROM base_conciliacao').fetchone()['c']
    conn.close()

    linhas = []
    for r in resultado:
        calc     = r['pontos_calculados'] or 0
        fortinet = r['pontos_fortinet']   or 0
        diff     = calc - fortinet
        status   = 'ok' if abs(diff) < 0.01 else ('acima' if diff > 0 else 'abaixo')
        linhas.append({**dict(r), 'diferenca': diff, 'status': status})

    return render_template('conciliacao.html', linhas=linhas, total_base=total_base)


# ── Admin ─────────────────────────────────────────────────────────────────────

@app.route('/admin/limpar-banco', methods=['POST'])
@login_required
def limpar_banco():
    if not ALLOW_DB_RESET:
        flash('Limpeza do banco desabilitada neste ambiente.', 'erro')
        return redirect(url_for('dashboard'))

    backup_database('before_reset')
    conn = get_db_connection()
    conn.execute('DELETE FROM pontos_utilizados')
    conn.execute('DELETE FROM base_conciliacao')
    conn.execute('DELETE FROM pontos_bolsao')
    try:
        conn.execute('DELETE FROM sqlite_sequence')
    except Exception:
        pass
    conn.commit()
    conn.close()
    flash('Banco de dados limpo com sucesso!', 'sucesso')
    return redirect(url_for('dashboard'))


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
