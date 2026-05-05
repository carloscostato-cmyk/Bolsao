import openpyxl
from openpyxl.utils import get_column_letter
import json

ARQUIVO = 'Controle de Licenças Fortinet (estudo) 1.xlsx'

wb = openpyxl.load_workbook(ARQUIVO, data_only=True)
wb_f = openpyxl.load_workbook(ARQUIVO)

for aba_nome in wb.sheetnames:
    ws = wb[aba_nome]
    ws_f = wb_f[aba_nome]
    print(f'\n{"="*120}')
    print(f'### ABA: "{aba_nome}"')
    print(f'### Dimensoes: {ws.dimensions}')
    print(f'### Linhas: {ws.max_row} | Colunas: {ws.max_column}')
    print(f'{"="*120}')
    
    # CABEÇALHOS (linha 1)
    print(f'\n--- CABECALHOS (Linha 1) ---')
    for col in range(1, ws.max_column+1):
        cell = ws.cell(row=1, column=col)
        print(f'  Col {col:2d} ({get_column_letter(col):2s}): "{cell.value}"')
    
    # VALORES MERGED CELLS
    print(f'\n--- CELULAS MESCLADAS ---')
    for mc in ws.merged_cells.ranges:
        print(f'  {mc}')
    
    # TODOS OS DADOS E FORMULAS
    print(f'\n--- DADOS COMPLETOS + FORMULAS ---')
    for row in range(1, ws.max_row+1):
        linha_tem_dado = False
        dados_linha = []
        for col in range(1, ws.max_column+1):
            cell = ws.cell(row=row, column=col)
            cell_f = ws_f.cell(row=row, column=col)
            val = cell.value
            formula = cell_f.value if cell_f.value and str(cell_f.value).startswith('=') else None
            if val is not None or formula is not None:
                linha_tem_dado = True
                dados_linha.append(f'    [{row:3d},{get_column_letter(col):2s}] Valor: {val} | Formula: {formula}')
        if linha_tem_dado:
            for d in dados_linha:
                print(d)
    print()