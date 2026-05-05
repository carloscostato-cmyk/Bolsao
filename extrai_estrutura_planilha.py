import openpyxl
import os

def main():
    # Caminho absoluto do arquivo Excel
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_path = os.path.join(base_dir, 'Controle de Licenças Fortinet (estudo) 1.xlsx')
    print(f"Tentando abrir: {excel_path}")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"Erro ao abrir a planilha: {e}")
        return

    print('\nAbas encontradas:')
    for sheet in wb.sheetnames:
        print(f'\n--- {sheet} ---')
        ws = wb[sheet]
        # Pega os nomes das colunas (primeira linha)
        try:
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            print('Colunas:', headers)
            # Mostra as 3 primeiras linhas de dados
            for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
                print(row)
        except Exception as e:
            print(f"Erro ao ler dados da aba '{sheet}': {e}")

if __name__ == "__main__":
    main()
